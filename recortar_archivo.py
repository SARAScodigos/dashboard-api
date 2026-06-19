#!/usr/bin/env python3
"""Recorta un archivo CSV o XLSX conservando su fila de encabezados."""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path


def crear_ruta_salida(entrada: Path) -> Path:
    """Genera un nombre de salida sin sobrescribir el archivo original."""
    return entrada.with_name(f"{entrada.stem}_recortado{entrada.suffix}")


def detectar_encoding(entrada: Path) -> str:
    """Detecta las codificaciones más habituales en CSV exportados desde Excel."""
    with entrada.open("rb") as archivo:
        muestra = archivo.read(128 * 1024)

    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            muestra.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue

    return "latin-1"


def recortar_csv(
    entrada: Path,
    salida: Path,
    cantidad: int,
    encoding: str,
) -> int:
    """Copia la cabecera y hasta `cantidad` registros usando memoria constante."""
    encoding_real = detectar_encoding(entrada) if encoding == "auto" else encoding

    with entrada.open("r", encoding=encoding_real, newline="") as archivo_entrada:
        muestra = archivo_entrada.read(64 * 1024)
        archivo_entrada.seek(0)

        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        except csv.Error:
            dialecto = csv.excel

        lector = csv.reader(archivo_entrada, dialecto)

        with salida.open("w", encoding=encoding_real, newline="") as archivo_salida:
            escritor = csv.writer(archivo_salida, dialecto)

            cabecera = next(lector, None)
            if cabecera is None:
                return 0

            escritor.writerow(cabecera)
            copiadas = 0

            for fila in lector:
                if copiadas >= cantidad:
                    break
                escritor.writerow(fila)
                copiadas += 1

    return copiadas


def recortar_xlsx(
    entrada: Path,
    salida: Path,
    cantidad: int,
    nombre_hoja: str | None,
) -> tuple[int, str]:
    """Copia valores de una hoja XLSX en modos de lectura/escritura eficientes."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Para archivos XLSX instala openpyxl con: pip install openpyxl"
        ) from error

    libro_entrada = load_workbook(entrada, read_only=True, data_only=False)

    if nombre_hoja:
        if nombre_hoja not in libro_entrada.sheetnames:
            disponibles = ", ".join(libro_entrada.sheetnames)
            libro_entrada.close()
            raise ValueError(
                f"La hoja '{nombre_hoja}' no existe. Disponibles: {disponibles}"
            )
        hoja_entrada = libro_entrada[nombre_hoja]
    else:
        hoja_entrada = libro_entrada.active

    libro_salida = Workbook(write_only=True)
    hoja_salida = libro_salida.create_sheet(hoja_entrada.title)
    copiadas = 0

    for indice, fila in enumerate(hoja_entrada.iter_rows(values_only=True)):
        # Índice 0: cabecera. Los demás son registros de datos.
        if indice > cantidad:
            break
        hoja_salida.append(fila)
        if indice > 0:
            copiadas += 1

    libro_salida.save(salida)
    nombre_resultado = hoja_entrada.title
    libro_entrada.close()
    return copiadas, nombre_resultado


def construir_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Crea una copia reducida de un CSV o XLSX. La cantidad indicada "
            "corresponde a filas de datos y no incluye la cabecera."
        )
    )
    parser.add_argument("entrada", type=Path, help="Archivo CSV o XLSX original")
    parser.add_argument(
        "filas",
        type=int,
        help="Número máximo de filas de datos que se copiarán",
    )
    parser.add_argument(
        "-o",
        "--salida",
        type=Path,
        help="Ruta de salida (por defecto: nombre_recortado.extensión)",
    )
    parser.add_argument(
        "--hoja",
        help="Nombre de la hoja para XLSX (por defecto se usa la hoja activa)",
    )
    parser.add_argument(
        "--encoding",
        default="auto",
        help="Codificación para CSV (por defecto: detección automática)",
    )
    return parser.parse_args()


def main() -> int:
    argumentos = construir_argumentos()
    entrada = argumentos.entrada.expanduser().resolve()
    salida = (argumentos.salida or crear_ruta_salida(entrada)).expanduser().resolve()

    if argumentos.filas < 0:
        print("Error: el número de filas no puede ser negativo.", file=sys.stderr)
        return 2
    if not entrada.is_file():
        print(f"Error: no se encontró el archivo: {entrada}", file=sys.stderr)
        return 2
    if entrada == salida:
        print("Error: la salida debe ser diferente del archivo original.", file=sys.stderr)
        return 2

    salida.parent.mkdir(parents=True, exist_ok=True)
    extension = entrada.suffix.lower()

    try:
        if extension == ".csv":
            copiadas = recortar_csv(
                entrada,
                salida,
                argumentos.filas,
                argumentos.encoding,
            )
            detalle = ""
        elif extension == ".xlsx":
            copiadas, hoja = recortar_xlsx(
                entrada,
                salida,
                argumentos.filas,
                argumentos.hoja,
            )
            detalle = f" de la hoja '{hoja}'"
        else:
            print("Error: solo se admiten archivos .csv y .xlsx.", file=sys.stderr)
            return 2
    except (OSError, UnicodeError, ValueError, RuntimeError, csv.Error) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    print(f"Listo: se copiaron {copiadas:,} filas{detalle}.")
    print(f"Archivo creado: {salida}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
